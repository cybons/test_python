import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo


def generate_filepath_with_timestamp(
    directory: str, base_name="temp", extension="xlsx", separator="_"
) -> str:
    # 現在のタイムスタンプを取得
    timestamp = datetime.now().strftime("%y%m%d%H%M%S")
    save_name = f"{base_name}{separator}{timestamp}.{extension}"
    save_path = os.path.join(directory, save_name)
    return save_path


def save_dfs_to_excel_with_tables(
    dfs: dict, directory: str, base_name="sheeted_data"
) -> str:
    """
    辞書形式で提供された複数のDataFrameを、シート別にテーブル形式でExcelに保存する関数

    Parameters:
    - dfs: dict - 保存するDataFrameの辞書（キーがシート名、値がDataFrame）
    - directory: str - ファイルを保存するディレクトリ

    Returns:
    - str: 保存したExcelファイルのパス
    """
    # 保存するExcelファイルのパスを生成
    file_path = generate_filepath_with_timestamp(directory, base_name)

    # 新しいワークブックを作成
    wb = Workbook()
    # デフォルトのシートを削除
    wb.remove(wb.active)

    # 各DataFrameをシートに書き込む
    for sheet_name, df in dfs.items():
        ws = wb.create_sheet(title=sheet_name)

        # DataFrameを行ごとに書き込み
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        # 1行目を固定
        ws.freeze_panes = "A2"

        # テーブル範囲を設定
        end_col = get_column_letter(df.shape[1])  # 列の数をアルファベットに変換
        table_range = f"A1:{end_col}{df.shape[0] + 1}"
        table = Table(displayName=f"Table_{sheet_name}", ref=table_range)

        # テーブルスタイルを設定
        style = TableStyleInfo(
            name="TableStyleMedium2",  # 薄いストライプスタイル
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,  # ストライプ表示
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

        # ヘッダー行（1行目）の文字色を白に変更
        for col in range(1, df.shape[1] + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(color="FFFFFF")  # 白色に設定

    # ワークブックを保存
    wb.save(file_path)
    return file_path
