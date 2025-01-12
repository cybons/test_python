import logging
from dataclasses import dataclass
from typing import Optional, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

from .constants import OPERATOR_MAPPING, ORG_HIERARCHY_X, ORG_HIERARCHY_Y


@dataclass
class FilterCondition:
    """フィルタリング条件を表すデータクラス

    Attributes:
        condition_id (str): 条件の一意識別子
        similarity_index (str): 類似度指標の名前
        operator (str): 演算子（例: '>', '<', '==', etc.）
        group_min_users (Optional[int]): グループの最小ユーザー数
        group_max_users (Optional[int]): グループの最大ユーザー数
        value (Union[float, int]): フィルタリングに使用する値
        description (str): 条件の説明
    """

    condition_id: str  # 条件の一意識別子
    similarity_index: str  # 類似度指標の名前
    operator: str  # 演算子（例: '>', '<', '==', etc.）
    group_min_users: Optional[int]  # グループの最小ユーザー数
    group_max_users: Optional[int]  # グループの最大ユーザー数
    value: Union[float, int]  # フィルタリングに使用する値
    description: str  # 条件の説明

    @classmethod
    def from_series(cls, series: pd.Series) -> "FilterCondition":
        """pandas.Seriesからフィルタリング条件を生成"""
        return cls(
            condition_id=series["Condition ID"],
            similarity_index=series["Similarity Index"],
            operator=series["Operator"],
            group_min_users=series["Group Min Users"]
            if pd.notna(series["Group Min Users"])
            else None,
            group_max_users=series["Group Max Users"]
            if pd.notna(series["Group Max Users"])
            else None,
            value=series["Value"],
            description=series["Description"],
        )


class OrganizationFilter:
    """組織の類似度データに対するフィルタリング処理を行うクラス"""

    def __init__(self, similarity_df: pd.DataFrame, conditions_path: str):
        """
        コンストラクタ

        Parameters:
            similarity_df (pd.DataFrame): 類似度計算結果のDataFrame
            conditions_path (str): フィルタリング条件を含むExcelファイルのパス
        """
        self.similarity_df = similarity_df.copy()
        self.conditions_path = conditions_path
        self.logger = logging.getLogger(__name__)

    def apply_filters(self) -> pd.DataFrame:
        """
        フィルタリング条件を適用し、結果のDataFrameを返す

        Returns:
            pd.DataFrame: フィルタリング結果を含むDataFrame
        """
        try:
            # 条件の読み込みとバリデーション
            conditions = self._load_and_validate_conditions()

            # フィルタリング用の列を初期化
            self._initialize_filter_columns()

            # 基本的なフィルタリング（ユーザー数が3人未満のペアを除外）
            self._apply_basic_filters()

            # フィルタリング対象のデータを取得
            filtered_df = self.similarity_df[~self.similarity_df["is_excluded"]].copy()

            # 各条件を適用
            for condition in conditions:
                self._apply_condition(filtered_df, condition)

            # 高類似度ペアに基づく除外フラグの設定
            self._set_exclude_flags()

            # needs_review フラグを設定
            self.similarity_df["needs_review"] = (
                ~self.similarity_df["is_excluded"]
                & ~self.similarity_df["is_high_similarity"]
            )

            return self.similarity_df

        except Exception as e:
            self.logger.error(f"Error during filtering: {str(e)}")
            raise

    def _load_and_validate_conditions(self) -> list[FilterCondition]:
        """
        Excelファイルからフィルタリング条件を読み込み、演算子のバリデーションを行います。
        複数の条件がある場合は、各条件を個別の行として処理します。

        Returns:
            list[FilterCondition]: フィルタリング条件のリスト
        """
        try:
            conditions_df = pd.read_excel(self.conditions_path)

            # 必須列の確認
            required_columns = {
                "Condition ID",
                "Similarity Index",
                "Operator",
                "Group Min Users",
                "Group Max Users",
                "Value",
                "Description",
            }
            missing_columns = required_columns - set(conditions_df.columns)
            if missing_columns:
                raise ValueError(f"Missing required columns: {missing_columns}")

            # 演算子のバリデーション
            invalid_operators = set(conditions_df["Operator"].dropna()) - set(
                OPERATOR_MAPPING.keys()
            )
            if invalid_operators:
                raise ValueError(f"Unsupported operators found: {invalid_operators}")

            # FilterConditionオブジェクトのリストを作成
            return [
                FilterCondition.from_series(row) for _, row in conditions_df.iterrows()
            ]

        except Exception as e:
            self.logger.error(f"Error loading conditions: {str(e)}")
            raise

    def _initialize_filter_columns(self) -> None:
        """フィルタリング用の列を初期化"""
        self.similarity_df["is_excluded"] = False  # 除外フラグの初期化
        self.similarity_df["is_high_similarity"] = False  # 高類似度フラグの初期化
        self.similarity_df["matched_condition"] = ""

    def _apply_basic_filters(self) -> None:
        """基本的なフィルタリングを適用（ユーザー数が3人未満のペアを除外）"""
        self.similarity_df.loc[
            (self.similarity_df["num_users_df1"] < 3)
            | (self.similarity_df["num_users_df2"] < 3),
            "is_excluded",
        ] = True  # ユーザー数が3未満の行に除外フラグを設定

    def _apply_condition(
        self, filtered_df: pd.DataFrame, condition: FilterCondition
    ) -> None:
        """
        単一の条件をDataFrameに適用

        Excelファイルからフィルタリング条件を読み込み、'is_excluded' と 'is_high_similarity' 列を追加します。
        'is_excluded' 列はフィルタリングせずに全てのペアに対して設定します。
        'matched_condition' 列には、'is_high_similarity=True' となった際にマッチした条件を記録します。

        Parameters:
            filtered_df (pd.DataFrame): フィルタリング対象のDataFrame
            condition (FilterCondition): 適用する条件
        """
        # 条件の初期化（全てTrueで開始）
        mask = pd.Series([True] * len(filtered_df), index=filtered_df.index)

        # グループの最小ユーザー数の条件
        if condition.group_min_users is not None:
            mask &= (filtered_df["num_users_df1"] >= condition.group_min_users) & (
                filtered_df["num_users_df2"] >= condition.group_min_users
            )

        # グループの最大ユーザー数の条件
        if condition.group_max_users is not None:
            mask &= (filtered_df["num_users_df1"] <= condition.group_max_users) & (
                filtered_df["num_users_df2"] <= condition.group_max_users
            )

        # 類似度指標の条件
        if condition.similarity_index is not None:
            op_func = OPERATOR_MAPPING[condition.operator]
            mask &= op_func(filtered_df[condition.similarity_index], condition.value)

        # 条件を満たす行のインデックスを取得
        matched_indices = filtered_df[mask].index

        # 高類似度フラグを設定
        self.similarity_df.loc[matched_indices, "is_high_similarity"] = True

        # matched_conditionsの更新
        for idx in matched_indices:
            self.similarity_df.at[idx, "matched_condition"] = condition.description

        filtered_df.drop(matched_indices, inplace=True)

    def _set_exclude_flags(self) -> None:
        """高類似度ペアが存在する場合、同じ組織名の他のペアをis_excluded=Trueに設定"""
        high_similarity_pairs = self.similarity_df[
            self.similarity_df["is_high_similarity"]
        ]

        if not high_similarity_pairs.empty:

            def set_flags(org_column: str):
                """指定された組織名列に基づいて除外フラグを設定するヘルパー関数

                Parameters:
                    org_column (str): 組織名の列名（例: 'df1_org_full_name'）
                """
                orgs_to_exclude = pd.unique(high_similarity_pairs[org_column])
                self.similarity_df.loc[
                    (self.similarity_df[org_column].isin(orgs_to_exclude))
                    & (~self.similarity_df.index.isin(high_similarity_indices)),
                    "is_excluded",
                ] = True

            high_similarity_indices = (
                high_similarity_pairs.index
            )  # 高類似度ペア自体のインデックスを取得

            # df1_org_full_nameに基づく除外フラグの設定
            set_flags(ORG_HIERARCHY_X)

            # df2_org_full_nameに基づく除外フラグの設定
            set_flags(ORG_HIERARCHY_Y)

    def export_to_excel(self, output_path: str):
        """フィルタリング結果をExcelファイルに出力する

        Parameters:
            output_path (str): 出力するExcelファイルのパス
        """
        # 新しいワークブックを作成
        wb = Workbook()

        # テーブルスタイルの定義（薄いスタイル）
        table_style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        # 'All_Data' シートを取得（デフォルトでアクティブなシート）
        ws_all = wb.active
        ws_all.title = "All_Data"

        # self.similarity_df を 'All_Data' シートに書き込む（ヘッダーを1行目に）
        for r_idx, row in enumerate(
            dataframe_to_rows(self.similarity_df, index=False, header=True), 1
        ):
            for c_idx, value in enumerate(row, 1):
                ws_all.cell(row=r_idx, column=c_idx, value=value)

        # テーブル範囲を定義
        max_row, max_col = self.similarity_df.shape
        table_ref = (
            f"A1:{get_column_letter(max_col)}{max_row + 1}"  # +1 はヘッダー行を含むため
        )

        # テーブルを作成してシートに追加
        tab_all = Table(
            displayName="All_Data_Table", ref=table_ref, tableStyleInfo=table_style
        )
        ws_all.add_table(tab_all)

        # needs_review_df を抽出
        needs_review_df = self.similarity_df[self.similarity_df["needs_review"]]

        # 'Needs_Review' シートを作成
        ws_review = wb.create_sheet(title="Needs_Review")

        if not needs_review_df.empty:
            # needs_review_df を 'Needs_Review' シートに書き込む（ヘッダーを1行目に）
            for r_idx, row in enumerate(
                dataframe_to_rows(needs_review_df, index=False, header=True), 1
            ):
                for c_idx, value in enumerate(row, 1):
                    ws_review.cell(row=r_idx, column=c_idx, value=value)

            # テーブル範囲を定義
            nr_max_row, nr_max_col = needs_review_df.shape
            table_ref_review = f"A1:{get_column_letter(nr_max_col)}{nr_max_row + 1}"  # +1 はヘッダー行を含むため

            # テーブルを作成してシートに追加
            tab_review = Table(
                displayName="Needs_Review_Table",
                ref=table_ref_review,
                tableStyleInfo=table_style,
            )
            ws_review.add_table(tab_review)
        else:
            # データがない場合、A1にメッセージを入力
            msg = "対象データはありませんでした"
            ws_review["A1"] = msg
            # メッセージを強調表示（フォントサイズを大きくするなど）
            ws_review["A1"].font = Font(bold=True, size=14)

        # 自動幅調整関数の定義
        def auto_adjust_column_width(sheet, dataframe):
            for idx, col in enumerate(dataframe.columns, 1):
                column_letter = get_column_letter(idx)
                try:
                    max_length = (
                        max(dataframe[col].astype(str).map(len).max(), len(col)) + 2
                    )  # 余白を追加
                except ValueError:
                    max_length = len(col) + 2
                sheet.column_dimensions[column_letter].width = max_length

        # 'All_Data' シートの列幅を調整
        auto_adjust_column_width(ws_all, self.similarity_df)

        # 'Needs_Review' シートの列幅を調整
        if not needs_review_df.empty:
            auto_adjust_column_width(ws_review, needs_review_df)
        else:
            # データがない場合、A列の幅をメッセージに合わせて調整
            ws_review.column_dimensions["A"].width = (
                len("対象データはありませんでした") + 2
            )

        # Excelファイルを保存
        wb.save(output_path)
